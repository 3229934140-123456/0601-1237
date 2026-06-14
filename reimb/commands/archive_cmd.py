from __future__ import annotations

import csv
import json
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config import (
    load_task_state, save_task_state, load_task_config,
    EXPORT_DIR, append_log,
)
from ..models import Receipt, TaskStatus, ExportRecord
from ..utils import format_amount, group_by_project
from ..commands.export_cmd import (
    export_to_excel as export_excel_full,
    export_to_csv as export_csv_full,
    _add_export_record, create_batch_id,
)


def _generate_risk_report(task_dir: Path, receipts: list[Receipt],
                          month_filter: Optional[str],
                          batch_id: Optional[str] = None,
                          operation: str = "archive_month") -> Path:
    export_dir = task_dir / EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_suffix = f"_{month_filter}" if month_filter else ""
    filepath = export_dir / f"风险说明{month_suffix}_{timestamp}.xlsx"

    wb = Workbook()

    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws = wb.active
    ws.title = "高风险项"
    headers = ["序号", "票据ID", "文件名", "来源路径", "日期", "金额",
               "员工", "项目", "风险等级", "风险原因", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    risk_count = {"高": 0, "中": 0, "低": 0}
    for i, receipt in enumerate(receipts, 1):
        if receipt.risk_level not in ("高", "中"):
            continue
        risk_count[receipt.risk_level] = risk_count.get(receipt.risk_level, 0) + 1
        fill = high_fill if receipt.risk_level == "高" else medium_fill
        row_data = [
            i, receipt.id, receipt.filename, receipt.source_path,
            receipt.date or "", format_amount(receipt.amount),
            receipt.employee or "", receipt.project or "",
            receipt.risk_level, receipt.risk_reason, receipt.description or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col <= 10:
                cell.fill = fill
        row_idx += 1

    ws2 = wb.create_sheet("重复票据")
    dup_headers = ["序号", "票据ID", "文件名", "重复与", "说明"]
    for col, h in enumerate(dup_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    dup_count = 0
    for i, receipt in enumerate(receipts, 1):
        if not receipt.is_duplicate:
            continue
        dup_count += 1
        row_data = [
            i, receipt.id, receipt.filename,
            receipt.duplicate_of or "", "文件哈希或内容高度相似",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
        row_idx += 1

    ws3 = wb.create_sheet("缺附件票据")
    miss_headers = ["序号", "票据ID", "文件名", "票据类型", "缺失附件", "说明"]
    for col, h in enumerate(miss_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    miss_count = 0
    for i, receipt in enumerate(receipts, 1):
        if not receipt.is_missing_attachment:
            continue
        miss_count += 1
        missing = ", ".join(receipt.missing_attachments or [])
        row_data = [
            i, receipt.id, receipt.filename, receipt.receipt_type,
            missing, f"按规则缺少 {len(receipt.missing_attachments or [])} 个附件",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
        row_idx += 1

    ws4 = wb.create_sheet("未识别票据")
    unrec_headers = ["序号", "票据ID", "文件名", "识别状态", "缺失字段"]
    for col, h in enumerate(unrec_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    unrec_count = 0
    for i, receipt in enumerate(receipts, 1):
        missing_fields = []
        if not receipt.date:
            missing_fields.append("日期")
        if not receipt.amount:
            missing_fields.append("金额")
        if not receipt.employee:
            missing_fields.append("员工")
        if not missing_fields and receipt.risk_level == "无":
            continue
        unrec_count += 1
        row_data = [
            i, receipt.id, receipt.filename, receipt.extraction_status,
            ", ".join(missing_fields),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
        row_idx += 1

    ws5 = wb.create_sheet("汇总")
    summary = [
        ["统计项", "数量"],
        ["票据总数", len(receipts)],
        ["高风险项", risk_count["高"]],
        ["中风险项", risk_count["中"]],
        ["重复票据", dup_count],
        ["缺失附件", miss_count],
        ["未识别/不完整", unrec_count],
    ]
    for r_idx, row in enumerate(summary, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    wb.save(filepath)

    record = ExportRecord(
        export_type="风险说明",
        format="excel",
        filepath=str(filepath),
        record_count=len(receipts),
        total_amount=sum(r.amount or 0 for r in receipts),
        month_filter=month_filter,
        operation=operation,
    )
    _add_export_record(task_dir, record, batch_id=batch_id)

    return filepath


def _generate_summary_report(task_dir: Path, receipts: list[Receipt],
                             month_filter: Optional[str],
                             batch_id: Optional[str] = None,
                             operation: str = "archive_month") -> Path:
    export_dir = task_dir / EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    config = load_task_config(task_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_suffix = f"_{month_filter}" if month_filter else ""
    filepath = export_dir / f"汇总报告{month_suffix}_{timestamp}.md"

    total_count = len(receipts)
    total_amount = sum(r.amount or 0 for r in receipts)
    high_count = sum(1 for r in receipts if r.risk_level == "高")
    medium_count = sum(1 for r in receipts if r.risk_level == "中")
    dup_count = sum(1 for r in receipts if r.is_duplicate)
    miss_count = sum(1 for r in receipts if r.is_missing_attachment)
    unrec_count = sum(1 for r in receipts if r.extraction_status in ("识别失败", "部分识别", "待处理"))
    mod_count = sum(1 for r in receipts if r.is_modified)

    project_keywords = config.get_project_keywords_dict()
    groups = group_by_project(receipts, config.project_list, project_keywords=project_keywords)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    month_desc = month_filter if month_filter else "全部"

    md_content = []
    md_content.append(f"# 报销材料汇总报告")
    md_content.append("")
    md_content.append(f"**生成时间**: {now_str}")
    md_content.append(f"**统计范围**: {month_desc}")
    md_content.append(f"**规则版本**: v{config.rule_version}")
    if batch_id:
        md_content.append(f"**批次号**: {batch_id}")
    md_content.append("")
    md_content.append("## 📊 总体概览")
    md_content.append("")
    md_content.append(f"| 指标 | 数值 |")
    md_content.append(f"|------|------|")
    md_content.append(f"| 票据总数 | {total_count} |")
    md_content.append(f"| 总金额 | {format_amount(total_amount)} |")
    md_content.append(f"| 高风险项 | {high_count} |")
    md_content.append(f"| 中风险项 | {medium_count} |")
    md_content.append(f"| 重复票据 | {dup_count} |")
    md_content.append(f"| 缺失附件 | {miss_count} |")
    md_content.append(f"| 识别不完整 | {unrec_count} |")
    md_content.append(f"| 人工修正 | {mod_count} |")
    md_content.append("")

    md_content.append("## 📁 项目分组汇总")
    md_content.append("")
    md_content.append("| 项目 | 票据数 | 总金额 | 高风险 | 中风险 |")
    md_content.append("|------|--------|--------|--------|--------|")
    for project, prs in sorted(groups.items()):
        pr_amount = sum(r.amount or 0 for r in prs)
        pr_high = sum(1 for r in prs if r.risk_level == "高")
        pr_med = sum(1 for r in prs if r.risk_level == "中")
        md_content.append(f"| {project} | {len(prs)} | {format_amount(pr_amount)} | {pr_high} | {pr_med} |")
    md_content.append("")

    md_content.append("## ⚠️  高风险明细摘要")
    md_content.append("")
    high_risk_receipts = [r for r in receipts if r.risk_level == "高"]
    if high_risk_receipts:
        md_content.append("| 文件名 | 日期 | 金额 | 员工 | 项目 | 原因 |")
        md_content.append("|--------|------|------|------|------|------|")
        for r in high_risk_receipts[:20]:
            md_content.append(
                f"| {r.filename} | {r.date or '-'} | {format_amount(r.amount)} | "
                f"{r.employee or '-'} | {r.project or '-'} | {r.risk_reason or '-'} |"
            )
        if len(high_risk_receipts) > 20:
            md_content.append(f"")
            md_content.append(f"*(仅显示前 20 条，共 {len(high_risk_receipts)} 条高风险项)*")
    else:
        md_content.append("✅ 本次没有高风险项")
    md_content.append("")

    md_content.append("## 📝 修改摘要")
    md_content.append("")
    modified_receipts = [r for r in receipts if r.is_modified]
    if modified_receipts:
        md_content.append("| 文件名 | 修改次数 | 最近修改 |")
        md_content.append("|--------|----------|----------|")
        for r in modified_receipts:
            last_ts = r.field_modifications[-1].timestamp if r.field_modifications else "-"
            md_content.append(f"| {r.filename} | {len(r.field_modifications)} | {last_ts} |")
    else:
        md_content.append("暂无人工修正记录")
    md_content.append("")
    md_content.append("---")
    md_content.append(f"*本报告由 AI 报销整理工具自动生成于 {now_str}*")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(md_content))

    record = ExportRecord(
        export_type="汇总报告",
        format="markdown",
        filepath=str(filepath),
        record_count=total_count,
        total_amount=total_amount,
        month_filter=month_filter,
        operation=operation,
    )
    _add_export_record(task_dir, record, batch_id=batch_id)

    return filepath


def preview_monthly_archive(task_dir: Path, month: Optional[str] = None) -> dict:
    """预览月度归档：只统计不生成文件"""
    state = load_task_state(task_dir)
    all_receipts = [Receipt.from_dict(r) for r in state.receipts]

    if month:
        target_months = [month]
    else:
        target_months = sorted({r.date[:7] for r in all_receipts if r.date})
        if not target_months:
            target_months = [None]

    previews = []
    for m in target_months:
        if m:
            month_receipts = [r for r in all_receipts if r.date and r.date.startswith(m)]
        else:
            month_receipts = list(all_receipts)

        if not month_receipts:
            continue

        total_amount = sum(r.amount or 0 for r in month_receipts)
        high_risk = sum(1 for r in month_receipts if r.risk_level == "高")
        medium_risk = sum(1 for r in month_receipts if r.risk_level == "中")
        duplicates = sum(1 for r in month_receipts if r.is_duplicate)
        missing_att = sum(1 for r in month_receipts if r.is_missing_attachment)
        modified = sum(1 for r in month_receipts if r.is_modified)

        file_list = [
            ("报销清单", "Excel/CSV"),
            ("风险说明", "Excel"),
            ("汇总报告", "Markdown"),
            ("归档包", "ZIP"),
        ]

        previews.append({
            "month": m,
            "receipt_count": len(month_receipts),
            "total_amount": total_amount,
            "high_risk": high_risk,
            "medium_risk": medium_risk,
            "duplicates": duplicates,
            "missing_attachments": missing_att,
            "modified": modified,
            "files_to_generate": file_list,
        })

    return {
        "preview_count": len(previews),
        "previews": previews,
        "total_receipts": len(all_receipts),
    }


def monthly_archive(task_dir: Path, month: Optional[str] = None,
                    fmt: str = "excel", create_zip: bool = True,
                    batch_id: Optional[str] = None) -> dict:
    state = load_task_state(task_dir)
    config = load_task_config(task_dir)
    all_receipts = [Receipt.from_dict(r) for r in state.receipts]

    if month:
        target_months = [month]
    else:
        target_months = sorted({r.date[:7] for r in all_receipts if r.date})
        if not target_months:
            target_months = [None]

    if batch_id is None:
        batch_id = create_batch_id(task_dir)

    archive_results = []
    all_generated_files: list[Path] = []

    for m in target_months:
        if m:
            month_receipts = [r for r in all_receipts if r.date and r.date.startswith(m)]
        else:
            month_receipts = list(all_receipts)

        if not month_receipts:
            continue

        if fmt == "csv":
            list_path = export_csv_full(task_dir, month_filter=m, batch_id=batch_id,
                                       operation="archive_month")
        else:
            list_path = export_excel_full(task_dir, month_filter=m, batch_id=batch_id,
                                         operation="archive_month")

        risk_path = _generate_risk_report(task_dir, month_receipts,
                                          month_filter=m, batch_id=batch_id,
                                          operation="archive_month")
        report_path = _generate_summary_report(task_dir, month_receipts,
                                                month_filter=m, batch_id=batch_id,
                                                operation="archive_month")

        record_count = len(month_receipts)
        total_amount = sum(r.amount or 0 for r in month_receipts)

        zip_path = None
        if create_zip:
            export_dir = task_dir / EXPORT_DIR
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            m_suffix = f"_{m}" if m else ""
            zip_path = export_dir / f"月度归档包{m_suffix}_{timestamp}.zip"
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in [list_path, risk_path, report_path]:
                    zf.write(f, arcname=f.name)

            zip_record = ExportRecord(
                export_type=f"{m}归档包" if m else "归档包",
                format="zip",
                filepath=str(zip_path.resolve()),
                record_count=record_count,
                total_amount=total_amount,
                month_filter=m,
                operation="archive_month",
            )
            _add_export_record(task_dir, zip_record, batch_id=batch_id)
            all_generated_files.append(zip_path)

        archive_results.append({
            "month": m,
            "record_count": record_count,
            "total_amount": total_amount,
            "list_path": list_path,
            "risk_path": risk_path,
            "report_path": report_path,
            "zip_path": zip_path,
        })
        all_generated_files.extend([list_path, risk_path, report_path])

    state = load_task_state(task_dir)
    state.status = TaskStatus.EXPORTED.value
    save_task_state(task_dir, state)

    month_list = ", ".join(m for m in target_months if m) or "全部"
    append_log(
        task_dir, "archive",
        f"月度归档完成 (批次 {batch_id}): {len(archive_results)}个月({month_list}), "
        f"生成 {len(all_generated_files)} 个文件"
    )

    return {
        "archive_count": len(archive_results),
        "archives": archive_results,
        "total_files": len(all_generated_files),
        "all_files": all_generated_files,
        "batch_id": batch_id,
    }
