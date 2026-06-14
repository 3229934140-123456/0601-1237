from __future__ import annotations

import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

from ..config import load_task_state, load_task_config, EXPORT_DIR, append_log
from ..models import Receipt, TaskStatus, ExportRecord
from ..utils import format_amount


def _add_export_record(task_dir: Path, export_record: ExportRecord) -> None:
    from ..config import load_task_state, save_task_state
    state = load_task_state(task_dir)
    state.export_records.append(export_record)
    state.status = TaskStatus.EXPORTED.value
    save_task_state(task_dir, state)


def export_to_csv(task_dir: Path, month_filter: Optional[str] = None) -> Path:
    state = load_task_state(task_dir)
    config = load_task_config(task_dir)
    export_dir = task_dir / EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    receipts = [Receipt.from_dict(r) for r in state.receipts]
    active_month = month_filter if month_filter else config.month_filter
    if active_month:
        receipts = [r for r in receipts if r.date and r.date.startswith(active_month)]

    total_amount = sum(r.amount or 0 for r in receipts)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_suffix = f"_{active_month}" if active_month else ""
    filepath = export_dir / f"报销清单{month_suffix}_{timestamp}.csv"

    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "序号", "票据ID", "来源路径", "文件名", "识别状态", "票据类型",
            "日期", "金额", "员工", "项目", "是否重复", "缺附件",
            "风险等级", "风险原因", "修改标记", "备注",
        ])
        for i, receipt in enumerate(receipts, 1):
            writer.writerow([
                i, receipt.id, receipt.source_path, receipt.filename,
                receipt.extraction_status, receipt.receipt_type,
                receipt.date or "", format_amount(receipt.amount),
                receipt.employee or "", receipt.project or "",
                "是" if receipt.is_duplicate else "否",
                "是" if receipt.is_missing_attachment else "否",
                receipt.risk_level, receipt.risk_reason,
                "已修改" if receipt.is_modified else "",
                receipt.description,
            ])

    record = ExportRecord(
        export_type="报销清单",
        format="csv",
        filepath=str(filepath),
        record_count=len(receipts),
        total_amount=total_amount,
        month_filter=active_month,
    )
    _add_export_record(task_dir, record)
    append_log(task_dir, "export", f"导出CSV: {filepath.name}, {len(receipts)}条, {format_amount(total_amount)}")
    return filepath


def export_to_excel(task_dir: Path, month_filter: Optional[str] = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return export_to_csv(task_dir, month_filter=month_filter)

    state = load_task_state(task_dir)
    config = load_task_config(task_dir)
    export_dir = task_dir / EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    receipts = [Receipt.from_dict(r) for r in state.receipts]
    active_month = month_filter if month_filter else config.month_filter
    if active_month:
        receipts = [r for r in receipts if r.date and r.date.startswith(active_month)]

    total_amount = sum(r.amount or 0 for r in receipts)

    wb = Workbook()

    headers = [
        "序号", "票据ID", "来源路径", "来源子目录", "文件名", "识别状态",
        "票据类型", "日期", "金额", "员工", "项目", "是否重复",
        "缺附件", "风险等级", "风险原因", "修改标记", "备注",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    mod_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    ws = wb.active
    ws.title = "报销清单"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, receipt in enumerate(receipts, 1):
        row_num = i + 1
        row_data = [
            i, receipt.id, receipt.source_path, receipt.source_subdir,
            receipt.filename, receipt.extraction_status,
            receipt.receipt_type, receipt.date or "",
            format_amount(receipt.amount),
            receipt.employee or "", receipt.project or "",
            "是" if receipt.is_duplicate else "否",
            "是" if receipt.is_missing_attachment else "否",
            receipt.risk_level, receipt.risk_reason,
            "已修改" if receipt.is_modified else "",
            receipt.description,
        ]
        fill = None
        if receipt.risk_level == "高":
            fill = high_fill
        elif receipt.risk_level == "中":
            fill = medium_fill
        elif receipt.is_modified:
            fill = mod_fill

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    col_widths = [8, 14, 40, 15, 20, 12, 12, 12, 12, 10, 12, 10, 10, 10, 25, 10, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    if state.groups:
        ws2 = wb.create_sheet("项目汇总")
        g_headers = ["项目", "票据数", "总金额", "高风险数", "中风险数"]
        for col, h in enumerate(g_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for project, r_list in state.groups.items():
            if active_month:
                r_list = [r for r in r_list if r.get("date", "").startswith(active_month)]
            if not r_list:
                continue
            pr = [Receipt.from_dict(r) for r in r_list]
            total = sum(r.amount or 0 for r in pr)
            hc = sum(1 for r in pr if r.risk_level == "高")
            mc = sum(1 for r in pr if r.risk_level == "中")
            ws2.cell(row=row_idx, column=1, value=project)
            ws2.cell(row=row_idx, column=2, value=len(pr))
            ws2.cell(row=row_idx, column=3, value=format_amount(total))
            ws2.cell(row=row_idx, column=4, value=hc)
            ws2.cell(row=row_idx, column=5, value=mc)
            row_idx += 1

    ws3 = wb.create_sheet("差异说明")
    d_headers = ["票据ID", "文件名", "差异项", "原始值", "说明"]
    for col, h in enumerate(d_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for receipt in receipts:
        if not receipt.date:
            ws3.cell(row=row_idx, column=1, value=receipt.id)
            ws3.cell(row=row_idx, column=2, value=receipt.filename)
            ws3.cell(row=row_idx, column=3, value="日期")
            ws3.cell(row=row_idx, column=4, value="(空)")
            ws3.cell(row=row_idx, column=5, value="未能识别日期信息")
            row_idx += 1
        if not receipt.amount:
            ws3.cell(row=row_idx, column=1, value=receipt.id)
            ws3.cell(row=row_idx, column=2, value=receipt.filename)
            ws3.cell(row=row_idx, column=3, value="金额")
            ws3.cell(row=row_idx, column=4, value="(空)")
            ws3.cell(row=row_idx, column=5, value="未能识别金额信息")
            row_idx += 1
        if receipt.is_duplicate:
            ws3.cell(row=row_idx, column=1, value=receipt.id)
            ws3.cell(row=row_idx, column=2, value=receipt.filename)
            ws3.cell(row=row_idx, column=3, value="重复")
            ws3.cell(row=row_idx, column=4, value=f"与{receipt.duplicate_of}重复")
            ws3.cell(row=row_idx, column=5, value="文件哈希或内容高度相似")
            row_idx += 1
        if receipt.is_missing_attachment:
            ws3.cell(row=row_idx, column=1, value=receipt.id)
            ws3.cell(row=row_idx, column=2, value=receipt.filename)
            ws3.cell(row=row_idx, column=3, value="缺附件")
            ws3.cell(row=row_idx, column=4, value=",".join(receipt.missing_attachments))
            ws3.cell(row=row_idx, column=5, value="缺少必要附件")
            row_idx += 1

    ws4 = wb.create_sheet("修改历史")
    m_headers = ["票据ID", "文件名", "字段", "旧值", "新值", "时间", "操作人"]
    for col, h in enumerate(m_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for receipt in receipts:
        for mod in receipt.field_modifications:
            mod_dict = mod.to_dict() if hasattr(mod, "to_dict") else mod
            ws4.cell(row=row_idx, column=1, value=receipt.id)
            ws4.cell(row=row_idx, column=2, value=receipt.filename)
            ws4.cell(row=row_idx, column=3, value=mod_dict.get("field", ""))
            ws4.cell(row=row_idx, column=4, value=str(mod_dict.get("old_value", "")))
            ws4.cell(row=row_idx, column=5, value=str(mod_dict.get("new_value", "")))
            ws4.cell(row=row_idx, column=6, value=mod_dict.get("timestamp", ""))
            ws4.cell(row=row_idx, column=7, value=mod_dict.get("operator", "财务人员"))
            row_idx += 1

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_suffix = f"_{active_month}" if active_month else ""
    filepath = export_dir / f"报销清单{month_suffix}_{timestamp}.xlsx"
    wb.save(filepath)

    record = ExportRecord(
        export_type="报销清单",
        format="excel",
        filepath=str(filepath),
        record_count=len(receipts),
        total_amount=total_amount,
        month_filter=active_month,
    )
    _add_export_record(task_dir, record)
    append_log(task_dir, "export", f"导出Excel: {filepath.name}, {len(receipts)}条, {format_amount(total_amount)}")
    return filepath


def generate_report(task_dir: Path, month_filter: Optional[str] = None) -> Path:
    state = load_task_state(task_dir)
    config = load_task_config(task_dir)
    export_dir = task_dir / EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)

    receipts = [Receipt.from_dict(r) for r in state.receipts]
    active_month = month_filter if month_filter else config.month_filter
    if active_month:
        receipts = [r for r in receipts if r.date and r.date.startswith(active_month)]

    total_amount = sum(r.amount or 0 for r in receipts)
    total_count = len(receipts)
    duplicate_count = sum(1 for r in receipts if r.is_duplicate)
    missing_count = sum(1 for r in receipts if r.is_missing_attachment)
    high_risk_count = sum(1 for r in receipts if r.risk_level == "高")
    medium_risk_count = sum(1 for r in receipts if r.risk_level == "中")
    modified_count = sum(1 for r in receipts if r.is_modified)

    type_summary = {}
    for r in receipts:
        type_summary[r.receipt_type] = type_summary.get(r.receipt_type, 0) + 1

    status_summary = {}
    for r in receipts:
        status_summary[r.extraction_status] = status_summary.get(r.extraction_status, 0) + 1

    project_summary = {}
    for project, r_list in state.groups.items():
        if active_month:
            r_list = [r for r in r_list if r.get("date", "").startswith(active_month)]
        if not r_list:
            continue
        pr = [Receipt.from_dict(r) for r in r_list]
        project_summary[project] = {
            "count": len(pr),
            "total": sum(r.amount or 0 for r in pr),
        }

    high_risk_items = [r for r in receipts if r.risk_level == "高"]

    month_label = f"（{active_month}）" if active_month else ""
    lines = [
        "=" * 60,
        f"         报销材料汇总报告{month_label}",
        "=" * 60,
        "",
        f"任务名称: {state.task_name}",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"任务状态: {state.status}",
        f"规则版本: v{config.rule_version}",
        f"金额预警阈值: {format_amount(config.amount_warning_threshold)}",
        "",
        "-" * 40,
        "  总体统计",
        "-" * 40,
        f"  票据总数:     {total_count}",
        f"  总金额:       {format_amount(total_amount)}",
        f"  重复票据:     {duplicate_count}",
        f"  缺少附件:     {missing_count}",
        f"  高风险项:     {high_risk_count}",
        f"  中风险项:     {medium_risk_count}",
        f"  人工修改:     {modified_count}",
        "",
        "-" * 40,
        "  识别状态分布",
        "-" * 40,
    ]
    for status, count in sorted(status_summary.items(), key=lambda x: -x[1]):
        lines.append(f"  {status}: {count}张")

    lines.extend(["", "-" * 40, "  票据类型分布", "-" * 40])
    for rtype, count in sorted(type_summary.items(), key=lambda x: -x[1]):
        lines.append(f"  {rtype}: {count}张")

    if project_summary:
        lines.extend(["", "-" * 40, "  项目归类统计", "-" * 40])
        for project, info in project_summary.items():
            lines.append(f"  {project}: {info['count']}张, 金额 {format_amount(info['total'])}")

    if high_risk_items:
        lines.extend(["", "-" * 40, "  ⚠ 高风险项明细", "-" * 40])
        for r in high_risk_items:
            lines.append(f"  [{r.filename}] {r.risk_reason}")

    recent_exports = state.export_records[-5:] if hasattr(state, "export_records") else []
    if recent_exports:
        lines.extend(["", "-" * 40, "  最近导出文件", "-" * 40])
        for er in reversed(recent_exports):
            er_dict = er.to_dict() if hasattr(er, "to_dict") else er
            fp = Path(er_dict.get("filepath", ""))
            lines.append(f"  {er_dict.get('timestamp', '')[:19]} [{er_dict.get('format','').upper()}] {fp.name}")

    lines.extend(["", "=" * 60])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_suffix = f"_{active_month}" if active_month else ""
    filepath = export_dir / f"汇总报告{month_suffix}_{timestamp}.txt"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    record = ExportRecord(
        export_type="汇总报告",
        format="txt",
        filepath=str(filepath),
        record_count=total_count,
        total_amount=total_amount,
        month_filter=active_month,
    )
    _add_export_record(task_dir, record)
    append_log(task_dir, "export", f"生成汇总报告: {filepath.name}")
    return filepath


def export_by_month(task_dir: Path, fmt: str = "excel") -> dict:
    state = load_task_state(task_dir)
    receipts = [Receipt.from_dict(r) for r in state.receipts]

    months = sorted({r.date[:7] for r in receipts if r.date})
    if not months:
        return {"exported": 0, "months": []}

    exported = []
    for month in months:
        if fmt.lower() == "csv":
            fp = export_to_csv(task_dir, month_filter=month)
        else:
            fp = export_to_excel(task_dir, month_filter=month)
        exported.append({"month": month, "filepath": str(fp)})

    return {"exported": len(exported), "months": months, "files": exported}
